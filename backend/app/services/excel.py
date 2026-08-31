import csv
import io
from datetime import datetime
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..schema import EXPORT_COLUMNS, EXPORT_PRESETS, IMPORT_HEADER_MAP, IMPORT_POSITIONAL, TAB_BY_KEY, columns_for_preset
from .candidates import coerce_value, from_import_row
from ..schema import rating_average
from .common import age_from

COMPUTED_COLUMNS = (
    ("usia", "USIA", lambda d: age_from(d.get("tanggal_lahir")) or d.get("usia") or ""),
    ("nilai_rata", "NILAI RATA-RATA", lambda d: rating_average(d) or ""),
)

COMPUTED_MAP = {
    c[0]: c[2] for c in COMPUTED_COLUMNS
}


def _get_export_columns_for_preset(preset: str) -> List[Tuple[str, str, Optional[any]]]:
    """Return list of (key, header_label, getter_fn_or_None) for requested preset."""
    preset_cols = columns_for_preset(preset)
    cols = []
    for key, label in preset_cols:
        if key in COMPUTED_MAP:
            cols.append((key, label, COMPUTED_MAP[key]))
        else:
            cols.append((key, label, None))
    # If preset is full, ensure computed fields like USIA and NILAI RATA-RATA are included if not present
    if preset == "full":
        existing_keys = {c[0] for c in cols}
        if "usia" not in existing_keys:
            cols.append(("usia", "USIA", COMPUTED_MAP["usia"]))
        if "nilai_rata" not in existing_keys:
            cols.append(("nilai_rata", "NILAI RATA-RATA", COMPUTED_MAP["nilai_rata"]))
    return cols


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------
def build_export_workbook(docs: Iterable[dict], scope: str, preset: str = "full") -> Tuple[io.BytesIO, str]:
    """Return (stream xlsx, nama file) with professional styling."""
    wb = Workbook()
    ws = wb.active
    stage = TAB_BY_KEY.get(scope)
    ws.title = (stage.label.upper() if stage and stage.predicate else "MASTER DATA")[:31]
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    export_cols = _get_export_columns_for_preset(preset)
    headers = [label for _, label, _ in export_cols]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row_fill_even = PatternFill(start_color="F8FAFC", fill_type="solid")
    row_fill_odd = PatternFill(fill_type=None)
    data_font = Font(name="Calibri", size=10, color="0F172A")
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=Side(border_style="thin", color="334155"), right=Side(border_style="thin", color="334155"), top=Side(border_style="medium", color="0F172A"), bottom=Side(border_style="medium", color="0F172A"))

    # Format Header Row
    ws.row_dimensions[1].height = 28
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    # Append data rows
    row_idx = 2
    for doc in docs:
        row_values = []
        for key, _, fn in export_cols:
            val = fn(doc) if fn else doc.get(key, "")
            row_values.append(val if val is not None else "")
        
        ws.append(row_values)
        ws.row_dimensions[row_idx].height = 22
        current_fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        
        for col_num, cell in enumerate(ws[row_idx], 1):
            cell.fill = current_fill
            cell.font = data_font
            cell.border = border_all
            
            # Format values nicely
            key = export_cols[col_num - 1][0]
            val_str = str(cell.value or "")
            if "tanggal" in key or "created_at" in key:
                cell.alignment = data_align_center
            elif key in ("nik", "no_hp", "usia", "nilai_wajah", "nilai_komunikasi", "nilai_kedisiplinan", "nilai_rata"):
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left

        row_idx += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            # Give weight to line breaks or average lengths
            lines = val_str.split("\n")
            line_len = max(len(l) for l in lines) if lines else 0
            if line_len > max_len:
                max_len = line_len
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"recruitment_{scope}_{preset}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return stream, filename


def build_export_csv(docs: Iterable[dict], scope: str, preset: str = "full") -> Tuple[io.BytesIO, str]:
    """Return (stream csv, nama file)."""
    export_cols = _get_export_columns_for_preset(preset)
    stream = io.StringIO()
    writer = csv.writer(stream)
    
    headers = [label for _, label, _ in export_cols]
    writer.writerow(headers)

    for doc in docs:
        row_values = []
        for key, _, fn in export_cols:
            val = fn(doc) if fn else doc.get(key, "")
            row_values.append(val if val is not None else "")
        writer.writerow(row_values)

    bytes_stream = io.BytesIO(stream.getvalue().encode("utf-8-sig"))  # utf-8-sig for Excel compatibility
    filename = f"recruitment_{scope}_{preset}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return bytes_stream, filename


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------
# Minimal jumlah sel baris pertama yang harus dikenali agar dianggap header.
# 2 (bukan 1) supaya baris data yang kebetulan berisi kata seperti "catatan"
# tidak salah dibaca sebagai header.
MIN_HEADER_MATCHES = 2


def _column_fields(header: List[str]) -> Optional[List[Optional[str]]]:
    """Petakan header sheet -> daftar key field. None kalau header tak dikenali."""
    matches = sum(1 for h in header if h in IMPORT_HEADER_MAP)
    if matches < MIN_HEADER_MATCHES:
        return None
    return [IMPORT_HEADER_MAP.get(h) for h in header]


def parse_workbook(content: bytes) -> List[dict]:
    """Baca file .xlsx -> daftar dict kandidat yang valid.

    Baris pertama dipakai sebagai header kalau dikenali; kalau tidak, kolom
    dibaca berurutan sesuai IMPORT_POSITIONAL di schema.py.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        return []

    header = [str(c).strip().lower() if c is not None else "" for c in first]
    col_fields = _column_fields(header)
    if col_fields is None:
        # Tidak ada header -> baca semua baris termasuk baris pertama.
        rows_iter = ws.iter_rows(values_only=True)
        col_fields = list(IMPORT_POSITIONAL)

    out: List[dict] = []
    for row in rows_iter:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        values = {}
        for idx, raw in enumerate(row):
            if idx >= len(col_fields):
                break
            key = col_fields[idx]
            if not key or raw is None:
                continue
            value = coerce_value(key, raw)
            if value is not None:
                values[key] = value
        doc = from_import_row(values)
        if doc is not None:
            out.append(doc)
    return out


def build_import_template_workbook() -> Tuple[io.BytesIO, str]:
    """Return (stream xlsx, nama file) template impor data kandidat."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TEMPLATE IMPORT"
    ws.views.sheetView[0].showGridLines = True

    template_cols = [
        ("NAMA", "Budi Santoso (WAJIB)"),
        ("NIK (KTP)", "3171012308950001 (16 digit / kosongkan jika belum ada NIK)"),
        ("EMAIL", "budi@example.com"),
        ("NO HP", "081234567890"),
        ("TANGGAL LAHIR", "1995-08-17 (Format YYYY-MM-DD)"),
        ("POSISI APPLY", "Staff HR"),
        ("RENCANA PENEMPATAN", "Kantor Pusat"),
        ("ALAMAT", "Jl. Merdeka No. 10"),
        ("PENDIDIKAN TERAKHIR", "D4/S1"),
        ("PENGALAMAN KERJA", "HR Admin 2 tahun"),
        ("PIC", "Wardah"),
        ("KETERANGAN", "Kandidat rekomendasi"),
    ]

    headers = [col[0] for col in template_cols]
    sample_row_1 = [col[1].split(" (")[0] for col in template_cols]
    sample_row_2 = [
        "Siti Rahma",
        "",  # Kosongkan NIK
        "siti@example.com",
        "085712345678",
        "1998-12-05",
        "Kasir",
        "Toserba Payaman",
        "Jl. Sudirman No. 45",
        "SMA/SMK",
        "Kasir Retail 1 tahun",
        "Wardah",
        "NIK belum dikumpulkan",
    ]

    ws.append(headers)
    ws.append(sample_row_1)
    ws.append(sample_row_2)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample_font = Font(name="Calibri", size=10, italic=False, color="334155")
    data_align_left = Alignment(horizontal="left", vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=Side(border_style="thin", color="334155"), right=Side(border_style="thin", color="334155"), top=Side(border_style="medium", color="0F172A"), bottom=Side(border_style="medium", color="0F172A"))

    ws.row_dimensions[1].height = 28
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    for r_idx in (2, 3):
        ws.row_dimensions[r_idx].height = 22
        for c_idx, cell in enumerate(ws[r_idx], 1):
            cell.font = sample_font
            cell.border = border_all
            if c_idx in (2, 4, 5):  # NIK, HP, Tanggal Lahir
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left

    # Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = "template_import_kandidat_rekapin.xlsx"
    return stream, filename
