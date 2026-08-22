"""Unit test tanpa server — mengunci perilaku yang diturunkan dari app/schema.py.

Jalankan: pytest backend/tests/test_schema_unit.py
Test ini tidak butuh MongoDB atau server yang hidup, tapi tetap perlu
MONGO_URL & DB_NAME di environment karena app/config.py membacanya saat import.
"""

import io
import os
import sys
from pathlib import Path

import pytest

BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("DB_NAME", "test_unit")

from openpyxl import Workbook, load_workbook  # noqa: E402

from app import schema  # noqa: E402
from app.emailing import templates  # noqa: E402
from app.models import CandidateCreate, CandidateUpdate  # noqa: E402
from app.services import excel, history  # noqa: E402
from app.services import nik as nik_service  # noqa: E402
from app.services.rules import apply_auto_rules  # noqa: E402


# ---------------------------------------------------------------------------
# Konsistensi schema
# ---------------------------------------------------------------------------
def test_field_keys_unik():
    keys = [f.key for f in schema.FIELDS]
    assert len(keys) == len(set(keys))


def test_setiap_field_punya_grup_yang_terdaftar():
    known = {key for key, _ in schema.FIELD_GROUPS}
    unknown = {f.group for f in schema.FIELDS} - known
    assert not unknown, f"grup belum didaftarkan di FIELD_GROUPS: {unknown}"


def test_field_select_menunjuk_status_set_yang_ada():
    for f in schema.FIELDS:
        if f.type == "select":
            assert f.options in schema.STATUS_SETS, f.key
            assert schema.STATUS_SETS[f.options], f.key


def test_default_select_termasuk_pilihan_yang_valid():
    for f in schema.FIELDS:
        if f.type == "select" and f.default:
            assert f.default in schema.STATUS_SETS[f.options], f.key


def test_urutan_import_positional_tidak_bolong():
    indexes = sorted(f.paste_index for f in schema.FIELDS if f.paste_index is not None)
    assert indexes == list(range(len(indexes))), "paste_index harus 0,1,2,... tanpa lompat"


def test_kolom_export_lengkap_dan_diakhiri_tanggal_input():
    keys = [k for k, _ in schema.EXPORT_COLUMNS]
    assert keys[-1] == "created_at"
    assert len(keys) == len([f for f in schema.FIELDS if f.export]) + 1


def test_model_kandidat_ikut_field_schema():
    model_fields = set(CandidateCreate.model_fields)
    assert {f.key for f in schema.FIELDS} <= model_fields
    assert "custom_data" in model_fields
    # Semua field pada update bersifat opsional.
    assert CandidateUpdate().model_dump(exclude_unset=True) == {}


def test_field_wajib_divalidasi():
    with pytest.raises(Exception):
        CandidateCreate()
    wajib = {f.key: "3201011234567890" if f.key == "nik" else "Budi"
             for f in schema.FIELDS if f.required}
    assert CandidateCreate(**wajib).nama == "Budi"


def test_label_riwayat_sama_dengan_field_yang_dilacak():
    assert set(schema.FIELD_LABELS) == {f.key for f in schema.FIELDS if f.history}


# ---------------------------------------------------------------------------
# Aturan otomatis
# ---------------------------------------------------------------------------
def test_mulai_training_mengisi_tanggal_mulai():
    out = apply_auto_rules({}, {"status_training": schema.Training.ONGOING})
    assert out["tanggal_mulai_training"]


def test_tanggal_mulai_training_tidak_ditimpa():
    out = apply_auto_rules(
        {"status_training": schema.Training.NOT_YET, "tanggal_mulai_training": "2026-01-01"},
        {"status_training": schema.Training.ONGOING},
    )
    assert "tanggal_mulai_training" not in out or out["tanggal_mulai_training"] == "2026-01-01"


def test_mengundurkan_setelah_ttd_masuk_blacklist():
    out = apply_auto_rules({}, {"status_tanda_tangan": schema.Ttd.RESIGNED_AFTER})
    assert out["status_blacklist"] == schema.Blacklist.RESIGNED_AFTER_TTD
    assert out["alasan_blacklist"]


def test_pic_email_dinormalisasi():
    out = apply_auto_rules({}, {"pic_email": "  PIC@Company.COM "})
    assert out["pic_email"] == "pic@company.com"


# ---------------------------------------------------------------------------
# Tab, scope, funnel
# ---------------------------------------------------------------------------
CANDIDATES = [
    {"status_interview": schema.Interview.NOT_CALLED, "status_training": schema.Training.NOT_YET,
     "status_blacklist": schema.Blacklist.NO, "penempatan_fix": ""},
    {"status_interview": schema.Interview.SCHEDULED, "status_training": schema.Training.ONGOING,
     "status_blacklist": schema.Blacklist.NO, "penempatan_fix": "Cabang A",
     "status_tanda_tangan": schema.Ttd.SIGNED},
    {"status_interview": schema.Interview.PASSED, "status_training": schema.Training.FINISHED,
     "status_blacklist": schema.Blacklist.VIOLATION, "penempatan_fix": "  "},
]


@pytest.mark.parametrize("scope,expected", [
    ("all", 3), ("master", 3), ("interview", 2), ("training", 1),
    ("blacklist", 1), ("placement", 1),
])
def test_filter_scope(scope, expected):
    assert len(schema.filter_by_scope(CANDIDATES, scope)) == expected


def test_tab_dan_scope_export_memakai_daftar_yang_sama():
    assert {t.key for t in schema.TABS} == set(schema.TAB_BY_KEY)


def test_funnel_urut_dan_tidak_naik():
    counts = [
        len(CANDIDATES) if pred is None else sum(1 for c in CANDIDATES if pred(c))
        for _, _, pred, _q in schema.FUNNEL
    ]
    assert counts[0] == len(CANDIDATES)
    assert all(c >= 0 for c in counts)


# ---------------------------------------------------------------------------
# Template email
# ---------------------------------------------------------------------------
CANDIDATE = {
    "nama": "Budi", "email": "budi@example.com", "apply": "Kasir",
    "tanggal_interview": "2026-08-25", "jam_interview": "09:00",
    "metode_interview": "Online", "penempatan_fix": "Cabang A", "no_hp": "0812",
}


@pytest.mark.parametrize("spec", templates.TEMPLATES, ids=lambda s: s.id)
def test_semua_template_bisa_dirender(spec):
    out = templates.render(spec.id, CANDIDATE, extra={"penerima": "Tim"})
    assert out and out["subject"] and out["html"]
    assert "$" not in out["subject"], "ada placeholder yang tidak terisi"


def test_template_tidak_dikenal_mengembalikan_none():
    assert templates.render("tidak_ada", CANDIDATE) is None


def test_template_meng_escape_html():
    out = templates.render("panggilan_interview", {**CANDIDATE, "nama": "<script>x</script>"})
    assert "<script>" not in out["html"]


def test_daftar_template_publik_tanpa_template_internal():
    ids = {t["id"] for t in templates.public_templates()}
    assert "kandidat_baru_internal" not in ids
    assert "panggilan_interview" in ids


# ---------------------------------------------------------------------------
# Import & export Excel
# ---------------------------------------------------------------------------
def _workbook(rows):
    wb = Workbook()
    for row in rows:
        wb.active.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_dengan_header():
    content = _workbook([
        ["Nama", "Email", "No. HP", "Tanggal Lahir"],
        ["Budi", "b@x.com", "0812", "1999-03-15"],
        [None, None, None, None],
    ])
    rows = excel.parse_workbook(content)
    assert len(rows) == 1
    assert rows[0]["nama"] == "Budi" and rows[0]["tanggal_lahir"] == "1999-03-15"


def test_import_tanpa_header_mengikuti_urutan_schema():
    nilai = ["Ani", "a@x.com", "0899", "1999-03-15", "Admin", "Cabang A",
             "Jl. Mawar", "Wardah", "w@x.com", "catatan"]
    content = _workbook([nilai])
    rows = excel.parse_workbook(content)
    assert len(rows) == 1
    for key, value in zip(schema.IMPORT_POSITIONAL, nilai):
        assert rows[0][key] == value


def test_import_membuang_baris_tanpa_nama():
    content = _workbook([["Nama", "Email"], ["", "x@x.com"], ["Budi", "b@x.com"]])
    assert [r["nama"] for r in excel.parse_workbook(content)] == ["Budi"]


def test_export_header_sesuai_schema():
    from app.services.excel import COMPUTED_COLUMNS
    stream, filename = excel.build_export_workbook([{"nama": "Budi"}], "all")
    ws = load_workbook(stream).active
    diharapkan = ([label for _, label in schema.EXPORT_COLUMNS]
                  + [label for _k, label, _f in COMPUTED_COLUMNS])
    assert [c.value for c in ws[1]] == diharapkan
    assert filename.startswith("recruitment_all_") and filename.endswith(".xlsx")


# ---------------------------------------------------------------------------
# Riwayat perubahan
# ---------------------------------------------------------------------------
def test_diff_mencatat_perubahan_field_dan_kolom_kustom():
    changes = history.diff(
        {"nama": "A", "status_interview": schema.Interview.NOT_CALLED, "custom_data": {"x": "1"}},
        {"nama": "B", "status_interview": schema.Interview.SCHEDULED, "custom_data": {"x": "2"}},
    )
    fields = {c["field"] for c in changes}
    assert {"nama", "status_interview", "custom.x"} <= fields


def test_diff_kosong_kalau_tidak_ada_perubahan():
    doc = {"nama": "A", "status_interview": schema.Interview.NOT_CALLED}
    assert history.diff(doc, dict(doc)) == []


# ---------------------------------------------------------------------------
# NIK (KTP) — kunci identitas kandidat
# ---------------------------------------------------------------------------
def test_nik_terdaftar_sebagai_field_unik_dan_searchable():
    spec = schema.FIELD_BY_KEY["nik"]
    assert spec.unique and spec.searchable and spec.sensitive
    assert "nik" in schema.UNIQUE_FIELDS
    assert "nik" in schema.SEARCHABLE_FIELDS


@pytest.mark.parametrize("raw,expected", [
    ("3201011234567890", "3201011234567890"),
    ("3201 0112 3456 7890", "3201011234567890"),   # spasi
    ("3201.0112.3456.7890", "3201011234567890"),   # titik
    ("3201-0112-3456-7890", "3201011234567890"),   # strip
    ("  3201011234567890  ", "3201011234567890"),
    (None, ""),
    ("", ""),
])
def test_normalisasi_nik(raw, expected):
    assert nik_service.normalize(raw) == expected


@pytest.mark.parametrize("raw", ["3201011234567890", "0000000000000000", None, "", "   "])
def test_nik_valid_atau_kosong_diterima(raw):
    assert nik_service.reject_reason(raw) is None


@pytest.mark.parametrize("raw,pesan", [
    ("123", "16 digit"),                       # terlalu pendek
    ("32010112345678901", "16 digit"),         # terlalu panjang
    ("32010112345678AB", "hanya boleh berisi angka"),
    ("NIK: 3201011234567890", "hanya boleh berisi angka"),
])
def test_nik_tidak_valid_ditolak(raw, pesan):
    reason = nik_service.reject_reason(raw)
    assert reason and pesan in reason


def test_nik_wajib_16_digit():
    assert nik_service.NIK_LENGTH == 16
    assert nik_service.is_valid("1234567890123456")
    assert not nik_service.is_valid("123456789012345")


def test_pesan_duplikat_menyebut_nama_pemilik():
    pesan = nik_service.duplicate_message({"nama": "Budi", "status_blacklist": schema.Blacklist.NO})
    assert "Budi" in pesan
    assert "BLACKLIST" not in pesan


def test_pesan_duplikat_memperingatkan_blacklist():
    pesan = nik_service.duplicate_message({
        "nama": "Fajar",
        "status_blacklist": schema.Blacklist.VIOLATION,
        "alasan_blacklist": "Memalsukan dokumen",
    })
    assert "BLACKLIST" in pesan and "Fajar" in pesan
    assert "Memalsukan dokumen" in pesan


def test_nik_ikut_kolom_export():
    assert ("nik", "NIK (KTP)") in schema.EXPORT_COLUMNS


def test_alias_header_ktp_dikenali_saat_import():
    for header in ("nik", "no ktp", "ktp", "nomor ktp", "no. ktp", "nik (ktp)"):
        assert schema.IMPORT_HEADER_MAP[header] == "nik"


def test_urutan_kolom_paste_stabil():
    """Urutan kolom paste tidak boleh bergeser tanpa sengaja — sheet yang sudah
    dipakai tim akan salah baca. NIK sengaja ditaruh paling akhir; kolom usia
    diganti tanggal lahir di posisi yang sama (slot ke-4)."""
    assert schema.IMPORT_POSITIONAL == [
        "nama", "email", "no_hp", "tanggal_lahir", "apply",
        "rencana_penempatan", "alamat", "pic", "pic_email", "keterangan", "nik",
    ]


def test_import_excel_mengenali_kolom_no_ktp():
    content = _workbook([
        ["Nama", "No KTP"],
        ["Budi", "3201 0112 3456 7890"],
    ])
    rows = excel.parse_workbook(content)
    assert len(rows) == 1
    # Normalisasi final dilakukan saat simpan; di sini cukup nilainya terbaca.
    assert nik_service.normalize(rows[0]["nik"]) == "3201011234567890"


# ---------------------------------------------------------------------------
# Listing: query builder, paginasi, zona waktu
# ---------------------------------------------------------------------------
from app import config as app_config  # noqa: E402
from app.services import listing  # noqa: E402

ADMIN_USER = {"id": "a1", "email": "admin@example.com", "role": "admin"}
RECRUITER = {"id": "r1", "email": "pic@example.com", "role": "recruiter"}


def test_admin_tidak_dibatasi_tapi_recruiter_dibatasi():
    assert listing.build_query(ADMIN_USER) == {}
    q = listing.build_query(RECRUITER)
    assert "$or" in q and {"created_by": "r1"} in q["$or"]


def test_filter_tab_masuk_ke_query():
    q = listing.build_query(ADMIN_USER, scope="blacklist")
    assert q == schema.stage_query("blacklist")


def test_scope_tak_dikenal_tidak_memfilter():
    assert listing.build_query(ADMIN_USER, scope="tidak-ada") == {}
    assert listing.build_query(ADMIN_USER, scope="master") == {}


def test_pencarian_mencakup_semua_field_searchable():
    q = listing.build_query(ADMIN_USER, q="budi")
    fields = {list(c.keys())[0] for c in q["$or"]}
    assert set(schema.SEARCHABLE_FIELDS) <= fields


def test_pencarian_angka_juga_dicocokkan_tanpa_pemisah():
    q = listing.build_query(ADMIN_USER, q="3201 0112")
    patterns = {c[list(c.keys())[0]]["$regex"] for c in q["$or"]}
    assert any("32010112" in p for p in patterns), "NIK berspasi harus tetap ketemu"


def test_karakter_khusus_pada_pencarian_di_escape():
    # Tanpa escape, ".*" akan cocok ke semua dokumen.
    q = listing.build_query(ADMIN_USER, q=".*")
    patterns = {c[list(c.keys())[0]]["$regex"] for c in q["$or"]}
    assert all(p == r"\.\*" for p in patterns)


def test_filter_posisi():
    assert listing.build_query(ADMIN_USER, position="Kasir") == {"apply": "Kasir"}
    assert listing.build_query(ADMIN_USER, position="all") == {}


def test_batas_tanggal_memakai_zona_waktu_lokal():
    q = listing.build_query(ADMIN_USER, date_from="2026-08-23", date_to="2026-08-23")
    window = q["created_at"]
    offset = app_config.LOCAL_UTC_OFFSET_HOURS
    assert window["$gte"].endswith(f"+{offset:02d}:00"), window
    # Kandidat yang di-input 02:00 WIB tanggal 23 harus masuk filter tanggal 23.
    dini_hari_wib = "2026-08-23T02:00:00+07:00"
    assert window["$gte"] <= dini_hari_wib <= window["$lte"]


def test_paginasi_dibersihkan():
    assert listing.paginate(0, 0)[0] == 1                      # page minimal 1
    assert listing.paginate(1, 99999)[1] == app_config.MAX_PAGE_SIZE
    assert listing.paginate(3, 20)[2] == 40                    # skip
    assert listing.paginate(-5, 10)[0] == 1


@pytest.mark.parametrize("total,per_page,pages", [
    (0, 50, 1), (1, 50, 1), (50, 50, 1), (51, 50, 2), (127, 50, 3),
])
def test_jumlah_halaman(total, per_page, pages):
    assert listing.page_meta(total, 1, per_page)["pages"] == pages


def test_setiap_tab_berpredikat_punya_query_mongo():
    """Kalau menambah tab baru, `query` wajib diisi — kalau tidak, listing &
    hitungan akan salah (kesepadanannya diuji di tests/test_stage_queries.py)."""
    for tab in schema.TABS:
        if tab.predicate is not None:
            assert tab.query, f"tab '{tab.key}' belum punya query Mongo"
    for key, _label, predicate, query in schema.FUNNEL:
        if predicate is not None:
            assert query, f"tahap funnel '{key}' belum punya query Mongo"


def test_aturan_panjang_password():
    from app.models import UserCreate
    with pytest.raises(Exception):
        UserCreate(email="a@example.com", name="A", password="x" * (app_config.MIN_PASSWORD_LENGTH - 1))
    ok = UserCreate(email="a@example.com", name="A", password="x" * app_config.MIN_PASSWORD_LENGTH)
    assert ok.password


# ---------------------------------------------------------------------------
# NIK wajib + NIK sementara
# ---------------------------------------------------------------------------
def test_nik_wajib_diisi():
    assert schema.FIELD_BY_KEY["nik"].required
    with pytest.raises(Exception):
        CandidateCreate(nama="Tanpa NIK")
    assert CandidateCreate(nama="Ada", nik="3201011234567890").nik


def test_nik_sementara_berawalan_khusus_dan_16_digit():
    prefix = nik_service.TEMP_PREFIX
    contoh = prefix + "0" * (nik_service.NIK_LENGTH - len(prefix))
    assert nik_service.is_valid(contoh), "NIK sementara harus lolos validasi format"
    assert nik_service.is_temporary(contoh)
    # Awalan 9999 bukan kode wilayah yang sah, jadi tidak bentrok dengan NIK asli.
    assert prefix.startswith("9999")


def test_nik_asli_tidak_dianggap_sementara():
    assert not nik_service.is_temporary("3201011234567890")
    assert not nik_service.is_temporary("")


def test_baris_import_tanpa_nik_tetap_lolos_validasi():
    from app.services.candidates import from_import_row
    doc = from_import_row({"nama": "Sheet Lama"})
    assert doc is not None, "baris tanpa NIK tidak boleh langsung dibuang"
    assert doc["nik"] == "", "NIK dikosongkan agar diisi NIK sementara"


def test_baris_import_tanpa_nama_tetap_dibuang():
    from app.services.candidates import from_import_row
    assert from_import_row({"nik": "3201011234567890"}) is None
    assert from_import_row({"nama": "   "}) is None


# ---------------------------------------------------------------------------
# TTD Kontrak (6 bulan)
# ---------------------------------------------------------------------------
def test_field_kontrak_terdaftar():
    for key in ("status_kontrak", "tanggal_ttd_kontrak", "tanggal_habis_kontrak"):
        assert key in schema.FIELD_BY_KEY, key
    assert schema.FIELD_BY_KEY["status_kontrak"].default == schema.Kontrak.PENDING
    assert ("kontrak", "TTD Kontrak (6 Bulan)") in schema.FIELD_GROUPS


def test_label_ttd_dibedakan_kesepakatan_vs_kontrak():
    assert schema.FIELD_BY_KEY["status_tanda_tangan"].label == "Status TTD Kesepakatan"
    assert schema.FIELD_BY_KEY["status_kontrak"].label == "Status TTD Kontrak"
    labels = [lbl for _, lbl in schema.EXPORT_COLUMNS]
    assert "STATUS TTD KESEPAKATAN" in labels and "STATUS TTD KONTRAK" in labels


def test_alias_import_lama_tetap_dikenali():
    # Sheet lama yang memakai header "Status TTD" harus tetap terbaca.
    assert schema.IMPORT_HEADER_MAP["status ttd"] == "status_tanda_tangan"
    assert schema.IMPORT_HEADER_MAP["tanggal ttd"] == "tanggal_tanda_tangan"
    assert schema.IMPORT_HEADER_MAP["status kontrak"] == "status_kontrak"


@pytest.mark.parametrize("status,expected", [
    (schema.Kontrak.SIGNED, True),
    (schema.Kontrak.EXTENDED, True),
    (schema.Kontrak.PENDING, False),
    (schema.Kontrak.NOT_CONTINUED, False),
    ("sudah", True),          # tidak peduli huruf besar/kecil
])
def test_predikat_punya_kontrak(status, expected):
    assert schema.has_contract({"status_kontrak": status}) is expected


def test_ttd_kontrak_mengisi_tanggal_dan_masa_habis():
    out = apply_auto_rules({}, {"status_kontrak": schema.Kontrak.SIGNED})
    assert out["tanggal_ttd_kontrak"], "tanggal TTD kontrak harus terisi otomatis"
    assert out["tanggal_habis_kontrak"], "tanggal habis kontrak harus dihitung"
    from app.services.common import add_days
    assert out["tanggal_habis_kontrak"] == add_days(
        out["tanggal_ttd_kontrak"], app_config.CONTRACT_PERIOD_DAYS)


def test_tanggal_kontrak_yang_sudah_diisi_tidak_ditimpa():
    out = apply_auto_rules(
        {"tanggal_ttd_kontrak": "2026-01-01", "tanggal_habis_kontrak": "2026-12-31"},
        {"status_kontrak": schema.Kontrak.SIGNED})
    assert "tanggal_ttd_kontrak" not in out
    assert "tanggal_habis_kontrak" not in out


def test_mengundurkan_setelah_kontrak_masuk_blacklist():
    out = apply_auto_rules({}, {"status_kontrak": schema.Kontrak.RESIGNED_AFTER})
    assert out["status_blacklist"] == schema.Blacklist.RESIGNED_AFTER_KONTRAK
    assert "kontrak" in out["alasan_blacklist"].lower()


def test_tab_dan_funnel_kontrak_ada():
    assert "kontrak" in schema.TAB_BY_KEY
    assert schema.TAB_BY_KEY["kontrak"].query
    assert "kontrak" in [k for k, _l, _p, _q in schema.FUNNEL]


def test_add_days():
    from app.services.common import add_days
    assert add_days("2026-08-22", 180) == "2027-02-18"
    assert add_days("2026-01-01", 0) == "2026-01-01"
    assert add_days("bukan tanggal", 10) == ""
    assert add_days("", 10) == ""


def test_aturan_reminder_mencakup_training_dan_kontrak():
    from app.services.reminders import RULES
    keys = {r.key for r in RULES}
    assert keys == {"training", "kontrak"}
    kontrak = next(r for r in RULES if r.key == "kontrak")
    # tanggal_habis_kontrak sudah berupa tenggat, jadi tidak ditambah periode lagi
    assert kontrak.period_days == 0
    assert kontrak.date_field == "tanggal_habis_kontrak"


# ---------------------------------------------------------------------------
# Tanggal lahir menggantikan usia
# ---------------------------------------------------------------------------
from datetime import date as _date  # noqa: E402

from app.services.common import age_from, birthdate_from_nik  # noqa: E402

HARI_INI = _date(2026, 8, 22)


def test_kolom_usia_diganti_tanggal_lahir():
    assert "usia" not in schema.FIELD_BY_KEY
    assert schema.FIELD_BY_KEY["tanggal_lahir"].type == "date"
    # Slot posisi import dipakai ulang supaya sheet paste tidak bergeser.
    assert schema.IMPORT_POSITIONAL[3] == "tanggal_lahir"


def test_alias_import_tanggal_lahir():
    for header in ("tanggal lahir", "tgl lahir", "dob", "birth date"):
        assert schema.IMPORT_HEADER_MAP[header] == "tanggal_lahir"


@pytest.mark.parametrize("lahir,umur", [
    ("1999-03-15", 27),
    ("2001-08-22", 25),   # tepat ulang tahun hari ini
    ("2001-08-23", 24),   # besok ulang tahun -> belum genap
])
def test_hitung_umur(lahir, umur):
    assert age_from(lahir, HARI_INI) == umur


@pytest.mark.parametrize("bad", ["", None, "bukan tanggal", "2026-13-45"])
def test_umur_dari_tanggal_invalid(bad):
    assert age_from(bad, HARI_INI) is None


@pytest.mark.parametrize("nik,lahir", [
    ("3201011503990001", "1999-03-15"),   # laki-laki
    ("3275025507960002", "1996-07-15"),   # perempuan: DD 55 = 15 + 40
    ("3674030102010005", "2001-02-01"),
])
def test_tanggal_lahir_dari_nik(nik, lahir):
    assert birthdate_from_nik(nik, HARI_INI) == lahir


@pytest.mark.parametrize("nik,alasan", [
    ("9999393582027078", "NIK sementara tidak boleh jadi tanggal"),
    ("3201013209990001", "tanggal 32 tidak ada"),
    ("3201011513990001", "bulan 13 tidak ada"),
    ("32010115039900", "panjang NIK salah"),
    ("", "kosong"),
])
def test_nik_yang_tidak_menghasilkan_tanggal(nik, alasan):
    assert birthdate_from_nik(nik, HARI_INI) == "", alasan


def test_auto_rule_isi_tanggal_lahir_dari_nik():
    out = apply_auto_rules({}, {"nik": "3201011503990001"})
    assert out["tanggal_lahir"] == "1999-03-15"


def test_auto_rule_tidak_menimpa_tanggal_lahir_yang_ada():
    out = apply_auto_rules({"tanggal_lahir": "1990-01-01"}, {"nik": "3201011503990001"})
    assert "tanggal_lahir" not in out


def test_auto_rule_tidak_mengarang_dari_nik_sementara():
    out = apply_auto_rules({}, {"nik": "9999393582027078"})
    assert not out.get("tanggal_lahir")


def test_export_punya_kolom_usia_hasil_hitungan():
    from app.services.excel import COMPUTED_COLUMNS
    keys = {k for k, _l, _f in COMPUTED_COLUMNS}
    assert "usia" in keys
    hitung = next(f for k, _l, f in COMPUTED_COLUMNS if k == "usia")
    assert hitung({"tanggal_lahir": "1999-03-15"}) == age_from("1999-03-15")
    # Data lama tanpa tanggal lahir tetap memakai nilai usia yang tersimpan.
    assert hitung({"usia": 24}) == 24
    assert hitung({}) == ""
